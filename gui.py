#!/usr/bin/env python3
"""
料号翻译工具 - 桌面版
双击运行，拖入周报，一键翻译。无需网络、无需浏览器。
"""

import os
import sys
import re
import shutil
import tempfile
import zipfile
import threading
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import xml.etree.ElementTree as ET

import openpyxl
import xlrd

# ── 内置 Mapping 路径 ─────────────────────────────────────────────
if getattr(sys, 'frozen', False):
    BASE_DIR = Path(sys._MEIPASS)
else:
    BASE_DIR = Path(__file__).parent
MAPPING_PATH = BASE_DIR / "mapping.xlsx"


def load_mapping(path):
    """读取 mapping，构建双向映射：无论输入新料号还是旧料号，都能翻译。
    同时读取 A/B 列（旧料号/新料号）和 J/K 列（Cross Reference/Spec Item）。
    重复项只保留第一条记录，避免被错误覆盖。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    mapping = {}  # {part_number: (old, new)}

    def add_pair(old_str, new_str):
        if old_str not in mapping:
            mapping[old_str] = (old_str, new_str)
        if new_str not in mapping:
            mapping[new_str] = (old_str, new_str)

    for row in ws.iter_rows(min_row=3, values_only=True):
        # A/B 列（旧料号 / 新料号）
        if row[0] and row[1]:
            old_v = str(row[0]).strip()
            new_v = str(row[1]).strip()
            if old_v and new_v and new_v.lower() != "nan":
                add_pair(old_v, new_v)

        # J/K 列（Cross Reference / Spec Item）
        if len(row) >= 11 and row[9] and row[10]:
            old_v = str(row[9]).strip()
            new_v = str(row[10]).strip()
            if old_v and new_v and new_v.lower() != "nan":
                add_pair(old_v, new_v)

    wb.close()
    return mapping


def is_part_number(value):
    if not value:
        return False
    return bool(re.match(r"^[AB][A-Z0-9]{3,}(\.[ABT])?$", value.strip()))


def translate_xlsx(report_path, output_path, progress_callback=None):
    NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    ET.register_namespace('', NS)
    ET.register_namespace('r',
        'http://schemas.openxmlformats.org/officeDocument/2006/relationships')

    if progress_callback:
        progress_callback("加载对照表...")
    mapping = load_mapping(MAPPING_PATH)

    if progress_callback:
        progress_callback("正在翻译...")
    tmpdir = Path(tempfile.mkdtemp(prefix='pn_trans_'))

    try:
        with zipfile.ZipFile(report_path, 'r') as z:
            z.extractall(tmpdir)

        ss_path = tmpdir / 'xl' / 'sharedStrings.xml'
        if not ss_path.exists():
            shutil.copy2(report_path, output_path)
            return {"translated": 0, "unchanged": 0, "not_found": 0}

        tree = ET.parse(ss_path)
        root = tree.getroot()
        stats = {"translated": 0, "unchanged": 0, "not_found": 0}

        for si in root.findall(f'{{{NS}}}si'):
            t_elem = si.find(f'{{{NS}}}t')
            if t_elem is None or t_elem.text is None:
                continue
            original = t_elem.text.strip()
            if not is_part_number(original):
                continue
            if original in mapping:
                old_val, new_val = mapping[original]
                if old_val != new_val:
                    # 统一输出格式：旧料号-新料号
                    t_elem.text = f"{old_val}-{new_val}"
                    t_elem.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')
                    stats["translated"] += 1
                else:
                    stats["unchanged"] += 1
            else:
                stats["not_found"] += 1

        tree.write(ss_path, xml_declaration=True, encoding='UTF-8')

        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for fpath in sorted(tmpdir.rglob('*')):
                if fpath.is_file():
                    arcname = str(fpath.relative_to(tmpdir))
                    zout.write(fpath, arcname)

        return stats
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

def translate_xls(report_path, output_path, progress_callback=None):
    """翻译 .xls 文件（Excel 97-2003 格式），输出为 .xlsx"""

    if progress_callback:
        progress_callback("加载对照表...")
    mapping = load_mapping(MAPPING_PATH)

    if progress_callback:
        progress_callback("正在读取文件...")

    # 用 xlrd 读取 .xls
    wb_in = xlrd.open_workbook(report_path)

    # 用 openpyxl 创建输出 .xlsx
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    stats = {"translated": 0, "unchanged": 0, "not_found": 0}
    total_sheets = wb_in.nsheets

    for si in range(total_sheets):
        ws_in = wb_in.sheet_by_index(si)

        if progress_callback:
            progress_callback(f"翻译中... ({si+1}/{total_sheets})")

        ws_out = wb_out.create_sheet(title=ws_in.name[:31])  # Excel 限制 31 字符

        for r in range(ws_in.nrows):
            for c in range(ws_in.ncols):
                cell_value = ws_in.cell_value(r, c)
                cell_type = ws_in.cell_type(r, c)

                if cell_type == xlrd.XL_CELL_TEXT and cell_value:
                    original = str(cell_value).strip()
                    if is_part_number(original) and original in mapping:
                        old_val, new_val = mapping[original]
                        if old_val != new_val:
                            ws_out.cell(row=r+1, column=c+1).value = f"{old_val}-{new_val}"
                            stats["translated"] += 1
                        else:
                            ws_out.cell(row=r+1, column=c+1).value = original
                            stats["unchanged"] += 1
                    else:
                        ws_out.cell(row=r+1, column=c+1).value = original
                elif cell_type == xlrd.XL_CELL_NUMBER:
                    ws_out.cell(row=r+1, column=c+1).value = cell_value
                elif cell_type == xlrd.XL_CELL_DATE:
                    import datetime
                    dt = xlrd.xldate_as_datetime(cell_value, wb_in.datemode)
                    ws_out.cell(row=r+1, column=c+1).value = dt
                elif cell_type == xlrd.XL_CELL_BOOLEAN:
                    ws_out.cell(row=r+1, column=c+1).value = bool(cell_value)
                elif cell_type == xlrd.XL_CELL_EMPTY:
                    pass  # 空单元格，不写入
                else:
                    # 其他类型转字符串
                    ws_out.cell(row=r+1, column=c+1).value = str(cell_value) if cell_value else None

    wb_out.save(output_path)
    return stats



def add_old_pn_column(report_path, output_path, header_name="老料号", progress_callback=None):
    """检测新料号占比 >90% 的列，左侧插入老料号列；
    文件中其他散落的新料号则翻译为「老料号-新料号」格式。"""

    if progress_callback:
        progress_callback("加载对照表...")
    mapping = load_mapping(MAPPING_PATH)

    def is_new_pn(val_str):
        """判断是否为新料号（在 mapping 中且 new != old）"""
        if val_str in mapping:
            old_v, new_v = mapping[val_str]
            return new_v == val_str and old_v != new_v
        return False

    def get_old_pn(val_str):
        """获取老料号"""
        if val_str in mapping:
            return mapping[val_str][0]
        return None

    is_xls = report_path.lower().endswith('.xls') and not report_path.lower().endswith('.xlsx')

    if is_xls:
        if progress_callback:
            progress_callback("正在读取 .xls 文件...")
        wb_in = xlrd.open_workbook(report_path)
        wb_out = openpyxl.Workbook()
        wb_out.remove(wb_out.active)
        total_sheets = wb_in.nsheets
        total_cols_added = 0
        total_translated = 0

        for si in range(total_sheets):
            ws_in = wb_in.sheet_by_index(si)
            if progress_callback:
                progress_callback(f"处理中... ({si+1}/{total_sheets})")

            ws_out = wb_out.create_sheet(title=ws_in.name[:31])

            # 先复制原数据
            for r in range(ws_in.nrows):
                for c in range(ws_in.ncols):
                    ct = ws_in.cell_type(r, c)
                    cv = ws_in.cell_value(r, c)
                    if ct == xlrd.XL_CELL_EMPTY:
                        continue
                    elif ct == xlrd.XL_CELL_NUMBER:
                        ws_out.cell(row=r+1, column=c+1).value = cv
                    elif ct == xlrd.XL_CELL_DATE:
                        import datetime
                        dt = xlrd.xldate_as_datetime(cv, wb_in.datemode)
                        ws_out.cell(row=r+1, column=c+1).value = dt
                    elif ct == xlrd.XL_CELL_BOOLEAN:
                        ws_out.cell(row=r+1, column=c+1).value = bool(cv)
                    else:
                        ws_out.cell(row=r+1, column=c+1).value = str(cv) if cv else None

            # 第一遍：找出新料号 >90% 的列
            high_pn_cols = set()
            for col_idx in range(ws_in.ncols):
                total = 0
                new_cnt = 0
                for row_idx in range(ws_in.nrows):
                    ct = ws_in.cell_type(row_idx, col_idx)
                    if ct != xlrd.XL_CELL_TEXT:
                        continue
                    cv = str(ws_in.cell_value(row_idx, col_idx)).strip()
                    if not cv:
                        continue
                    total += 1
                    if is_new_pn(cv):
                        new_cnt += 1
                if total > 0 and (new_cnt / total) > 0.9:
                    high_pn_cols.add(col_idx)

            total_cols_added += len(high_pn_cols)

            # 从右往左插入老料号列
            sorted_cols = sorted(high_pn_cols, reverse=True)
            for col_idx in sorted_cols:
                insert_pos = col_idx + 1  # 1-based
                ws_out.insert_cols(insert_pos)
                ws_out.cell(row=1, column=insert_pos).value = header_name
                for row_idx in range(ws_in.nrows):
                    src_cell = ws_out.cell(row=row_idx+1, column=insert_pos+1)
                    src_val = src_cell.value
                    if src_val is None:
                        continue
                    src_str = str(src_val).strip()
                    if not src_str:
                        continue
                    old_pn = get_old_pn(src_str)
                    if old_pn:
                        ws_out.cell(row=row_idx+1, column=insert_pos).value = old_pn
                    else:
                        ws_out.cell(row=row_idx+1, column=insert_pos).value = src_str

            # 构建「已处理列」集合（新列 + 原始列）
            processed_cols = set()
            shift = 0
            for orig_col in sorted(high_pn_cols):
                processed_cols.add(orig_col + shift + 1)      # 新插入的老料号列
                processed_cols.add(orig_col + shift + 2)      # 原始列
                shift += 1

            # 第二遍：其他列中的新料号 → "老料号-新料号"
            max_col_out = ws_out.max_column
            for col_idx in range(1, max_col_out + 1):
                if col_idx in processed_cols:
                    continue
                for row_idx in range(1, ws_in.nrows + 1):
                    cell = ws_out.cell(row=row_idx, column=col_idx)
                    cv = cell.value
                    if cv is None:
                        continue
                    cv_str = str(cv).strip()
                    if not cv_str:
                        continue
                    if is_new_pn(cv_str):
                        old_v, new_v = mapping[cv_str]
                        cell.value = f"{old_v}-{new_v}"
                        total_translated += 1

        wb_out.save(output_path)
        return {"sheets": total_sheets, "columns_added": total_cols_added, "translated": total_translated}

    else:
        # .xlsx: 用 openpyxl
        if progress_callback:
            progress_callback("正在读取 .xlsx 文件...")
        wb = openpyxl.load_workbook(report_path)
        total_sheets = len(wb.worksheets)
        total_cols_added = 0
        total_translated = 0

        for si, ws in enumerate(wb.worksheets):
            if progress_callback:
                progress_callback(f"处理中... ({si+1}/{total_sheets})")

            max_col = ws.max_column
            max_row = ws.max_row

            # 第一遍：找出新料号 >90% 的列
            high_pn_cols = set()
            for col_idx in range(1, max_col + 1):
                total = 0
                new_cnt = 0
                for row_idx in range(1, max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cv = cell.value
                    if cv is None:
                        continue
                    cv_str = str(cv).strip()
                    if not cv_str:
                        continue
                    total += 1
                    if is_new_pn(cv_str):
                        new_cnt += 1
                if total > 0 and (new_cnt / total) > 0.9:
                    high_pn_cols.add(col_idx)

            total_cols_added += len(high_pn_cols)

            # 从右往左插入老料号列
            sorted_cols = sorted(high_pn_cols, reverse=True)
            for col_idx in sorted_cols:
                ws.insert_cols(col_idx)
                ws.cell(row=1, column=col_idx).value = header_name
                for row_idx in range(2, max_row + 1):
                    src_cell = ws.cell(row=row_idx, column=col_idx + 1)
                    src_val = src_cell.value
                    if src_val is None:
                        continue
                    src_str = str(src_val).strip()
                    if not src_str:
                        continue
                    old_pn = get_old_pn(src_str)
                    if old_pn:
                        ws.cell(row=row_idx, column=col_idx).value = old_pn
                    else:
                        ws.cell(row=row_idx, column=col_idx).value = src_str

            # 构建「已处理列」集合
            processed_cols = set()
            shift = 0
            for orig_col in sorted(high_pn_cols):
                processed_cols.add(orig_col + shift)      # 新插入的老料号列
                processed_cols.add(orig_col + shift + 1)  # 原始列
                shift += 1

            # 第二遍：其他列中的新料号 → "老料号-新料号"
            new_max_col = ws.max_column
            for col_idx in range(1, new_max_col + 1):
                if col_idx in processed_cols:
                    continue
                for row_idx in range(1, max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cv = cell.value
                    if cv is None:
                        continue
                    cv_str = str(cv).strip()
                    if not cv_str:
                        continue
                    if is_new_pn(cv_str):
                        old_v, new_v = mapping[cv_str]
                        cell.value = f"{old_v}-{new_v}"
                        total_translated += 1

        wb.save(output_path)
        return {"sheets": total_sheets, "columns_added": total_cols_added, "translated": total_translated}


# ── GUI ────────────────────────────────────────────────────────────
class ColumnNameDialog:
    """弹窗输入新增列的第一行名称"""
    def __init__(self, parent):
        self.result = None
        self.top = tk.Toplevel(parent)
        self.top.title("输入列名")
        self.top.resizable(False, False)
        self.top.configure(bg="#f5f5f7")
        self.top.transient(parent)
        self.top.grab_set()

        # 居中
        self.top.update_idletasks()
        w, h = 340, 140
        sw = parent.winfo_screenwidth()
        sh = parent.winfo_screenheight()
        x = (sw - w) // 2
        y = (sh - h) // 2
        self.top.geometry(f"{w}x{h}+{x}+{y}")

        label = tk.Label(self.top, text="请输入新增列的第一行名称：",
                         font=("Microsoft YaHei UI", 11),
                         fg="#1d1d1f", bg="#f5f5f7")
        label.pack(pady=(20, 8))

        self.entry = tk.Entry(self.top, font=("Microsoft YaHei UI", 12),
                              width=24, justify="center")
        self.entry.insert(0, "老料号")
        self.entry.pack(pady=(0, 12))
        self.entry.select_range(0, "end")
        self.entry.focus_set()

        btn_frame = tk.Frame(self.top, bg="#f5f5f7")
        btn_frame.pack()

        cancel_btn = tk.Button(btn_frame, text="取消",
                               font=("Microsoft YaHei UI", 10),
                               bg="#e8e8ed", fg="#1d1d1f",
                               relief="flat", padx=16, pady=4,
                               command=self._cancel)
        cancel_btn.pack(side="left", padx=(0, 8))

        ok_btn = tk.Button(btn_frame, text="确定",
                           font=("Microsoft YaHei UI", 10, "bold"),
                           bg="#0071e3", fg="white",
                           relief="flat", padx=16, pady=4,
                           command=self._ok)
        ok_btn.pack(side="left")

        self.top.bind("<Return>", lambda e: self._ok())
        self.top.bind("<Escape>", lambda e: self._cancel())

    def _ok(self):
        self.result = self.entry.get()
        self.top.destroy()

    def _cancel(self):
        self.result = None
        self.top.destroy()


class TranslatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("料号翻译工具")
        self.root.geometry("520x480")
        self.root.resizable(False, False)

        # 设置窗口图标
        icon_path = BASE_DIR / "icons" / "icon.ico"
        try:
            self.root.iconbitmap(default=str(icon_path))
        except Exception:
            pass

        self.root.update_idletasks()
        w, h = 520, 480
        sw = root.winfo_screenwidth()
        sh = root.winfo_screenheight()
        x = (sw - w) // 2
        y = (sh - h) // 2
        self.root.geometry(f"{w}x{h}+{x}+{y}")

        self.bg = "#f5f5f7"
        self.card = "#ffffff"
        self.accent = "#0071e3"
        self.text = "#1d1d1f"
        self.sub = "#6e6e73"

        self.root.configure(bg=self.bg)
        self.file_path = None
        self._build_ui()

    def _build_ui(self):
        title = tk.Label(self.root, text="📦 料号翻译工具",
                         font=("Microsoft YaHei UI", 20, "bold"),
                         fg=self.text, bg=self.bg)
        title.pack(pady=(30, 4))

        version = tk.Label(self.root, text="V1.5",
                           font=("Microsoft YaHei UI", 9),
                           fg=self.sub, bg=self.bg)
        version.pack(pady=(0, 6))

        subtitle = tk.Label(self.root,
                            text="拖入或选择你需要的文件，一键翻译为【旧料号-新料号】格式\n靶向治疗，拒绝内耗",
                            font=("Microsoft YaHei UI", 11),
                            fg=self.sub, bg=self.bg)
        subtitle.pack(pady=(0, 20))

        card = tk.Frame(self.root, bg=self.card, bd=0,
                        highlightbackground="#d2d2d7",
                        highlightthickness=1, padx=24, pady=24)
        card.pack(fill="both", padx=30, pady=(0, 10))

        self.drop_frame = tk.Frame(card, bg=self.card, bd=2, relief="groove")
        self.drop_frame.pack(fill="both", expand=True, ipady=30)

        self.drop_label = tk.Label(self.drop_frame,
                                   text="点击选择你需要的文件",
                                   font=("Microsoft YaHei UI", 12),
                                   fg=self.sub, bg=self.card, justify="center")
        self.drop_label.pack(expand=True)

        self.file_label = tk.Label(self.drop_frame, text="",
                                   font=("Microsoft YaHei UI", 11, "bold"),
                                   fg=self.accent, bg=self.card)
        self.file_label.pack()

        self.drop_frame.bind("<Button-1>", lambda e: self._select_file())
        self.drop_label.bind("<Button-1>", lambda e: self._select_file())
        # 拖拽功能仅 macOS 支持，Windows 自动降级为按钮选择
        pass

        # ── 选择文件按钮（独立一行） ──
        select_frame = tk.Frame(card, bg=self.card)
        select_frame.pack(fill="x", pady=(16, 8))

        self.select_btn = tk.Button(select_frame, text="选择文件",
                                    font=("Microsoft YaHei UI", 11),
                                    bg="#e8e8ed", fg=self.text,
                                    relief="flat", padx=20, pady=6,
                                    command=self._select_file)
        self.select_btn.pack()

        # ── 两个功能按钮（独立一行，等宽排列） ──
        action_frame = tk.Frame(card, bg=self.card)
        action_frame.pack(fill="x", pady=(0, 0))
        action_frame.grid_columnconfigure(0, weight=1, uniform="action")
        action_frame.grid_columnconfigure(1, weight=1, uniform="action")

        self.add_col_btn = tk.Button(action_frame, text="A.增一列老料号",
                                     font=("Microsoft YaHei UI", 12, "bold"),
                                     bg="#34c759", fg="white",
                                     activebackground="#28a745",
                                     activeforeground="white",
                                     disabledforeground="white",
                                     relief="flat", padx=20, pady=10,
                                     borderwidth=0,
                                     highlightthickness=0,
                                     state="disabled",
                                     command=self._add_old_column)
        self.add_col_btn.grid(row=0, column=0, sticky="ew", padx=(0, 4))

        self.translate_btn = tk.Button(action_frame, text="B.原单元格翻译",
                                       font=("Microsoft YaHei UI", 12, "bold"),
                                       bg=self.accent, fg="white",
                                       activebackground="#0062c4",
                                       activeforeground="white",
                                       disabledforeground="white",
                                       relief="flat", padx=20, pady=10,
                                       borderwidth=0,
                                       highlightthickness=0,
                                       state="disabled",
                                       command=self._translate)
        self.translate_btn.grid(row=0, column=1, sticky="ew", padx=(4, 0))

        # 署名行
        signature = tk.Label(card,
                            text="ROCKYCENTER PRODUCTION",
                            font=("Microsoft YaHei UI", 8),
                            fg="#c7c7cc", bg=self.card,
                            justify="center")
        signature.pack(side="bottom", pady=(12, 0))

        self.progress = ttk.Progressbar(card, mode="indeterminate", length=400)

        self.status_label = tk.Label(card, text="",
                                     font=("Microsoft YaHei UI", 10),
                                     fg=self.sub, bg=self.card)

    def _select_file(self):
        path = filedialog.askopenfilename(
            title="选择周报文件",
            filetypes=[("Excel 文件", "*.xlsx;*.xls"), ("所有文件", "*.*")]
        )
        if path:
            self.file_path = path
            self.file_label.config(text=f"✅  {os.path.basename(path)}")
            self.translate_btn.config(state="normal")
            self.add_col_btn.config(state="normal")

    def _add_old_column(self):
        if not self.file_path:
            return

        # 弹窗让用户输入列名
        dialog = ColumnNameDialog(self.root)
        self.root.wait_window(dialog.top)
        header_name = dialog.result
        if header_name is None:
            return
        header_name = header_name.strip() or "老料号"

        default_out = self.file_path.rsplit('.', 1)[0] + '_with_oldpn.xlsx'
        output_path = filedialog.asksaveasfilename(
            title="保存结果",
            defaultextension=".xlsx",
            initialfile=os.path.basename(default_out),
            filetypes=[("Excel 文件", "*.xlsx")]
        )
        if not output_path:
            return

        self.translate_btn.config(state="disabled")
        self.add_col_btn.config(state="disabled", text="处理中...")
        self.select_btn.config(state="disabled")
        self.progress.pack(pady=(12, 0))
        self.progress.start()

        def done(result):
            self.root.after(0, lambda: self._on_add_col_done(result, output_path))

        def run():
            try:
                stats = add_old_pn_column(self.file_path, output_path,
                                          header_name=header_name,
                                          progress_callback=lambda msg: None)
                done(stats)
            except Exception as e:
                self.root.after(0, lambda: self._on_error(str(e)))

        threading.Thread(target=run, daemon=True).start()

    def _on_add_col_done(self, result, output_path):
        self.progress.stop()
        self.progress.pack_forget()
        self.translate_btn.config(state="normal")
        self.add_col_btn.config(state="normal", text="A.增一列老料号")
        self.select_btn.config(state="normal")
        sheets = result.get("sheets", 0)
        cols = result.get("columns_added", 0)
        msg = f"完成！\n\n处理 Sheet 数：{sheets}\n新增老料号列：{cols} 列\n\n文件保存至：\n{output_path}"
        messagebox.showinfo("处理完成", msg)

    def _on_drop(self, event):
        data = event.data
        if data:
            path = data.strip().strip('{').strip('}')
            if os.path.isfile(path) and (path.lower().endswith('.xlsx') or path.lower().endswith('.xls')):
                self.file_path = path
                self.file_label.config(text=f"✅  {os.path.basename(path)}")
                self.translate_btn.config(state="normal")
                self.add_col_btn.config(state="normal")

    def _translate(self):
        if not self.file_path:
            return
        if self.file_path.lower().endswith('.xls') and not self.file_path.lower().endswith('.xlsx'):
            base = self.file_path.rsplit('.', 1)[0]
            default_out = base + '_translated.xlsx'
        else:
            default_out = self.file_path.replace('.xlsx', '_translated.xlsx')
        output_path = filedialog.asksaveasfilename(
            title="保存翻译结果",
            defaultextension=".xlsx",
            initialfile=os.path.basename(default_out),
            filetypes=[("Excel 文件", "*.xlsx")]
        )
        if not output_path:
            return

        self.translate_btn.config(state="disabled", text="翻译中...")
        self.add_col_btn.config(state="disabled")
        self.select_btn.config(state="disabled")
        self.progress.pack(pady=(12, 0))
        self.progress.start()

        def done(stats):
            self.root.after(0, lambda: self._on_done(stats, output_path))

        def run():
            try:
                is_xls = self.file_path.lower().endswith('.xls') and not self.file_path.lower().endswith('.xlsx')
                if is_xls:
                    stats = translate_xls(self.file_path, output_path, progress_callback=lambda msg: None)
                else:
                    stats = translate_xlsx(self.file_path, output_path)
                done(stats)
            except Exception as e:
                self.root.after(0, lambda: self._on_error(str(e)))

        threading.Thread(target=run, daemon=True).start()

    def _on_done(self, stats, output_path):
        self.progress.stop()
        self.progress.pack_forget()
        self.translate_btn.config(state="normal", text="B.原单元格翻译")
        self.add_col_btn.config(state="normal", text="A.增一列老料号")
        self.select_btn.config(state="normal")

        t = stats.get("translated", 0)
        nf = stats.get("not_found", 0)

        msg = f"翻译完成！\n\n✅ 已翻译：{t} 个\n"
        if nf > 0:
            msg += f"⚠️ 未找到映射：{nf} 个（其他辅助 Sheet 中的旧编号，不影响阅读）\n"
        msg += f"\n文件已保存至：\n{output_path}"
        messagebox.showinfo("翻译完成", msg)

    def _on_error(self, err_msg):
        self.progress.stop()
        self.progress.pack_forget()
        self.translate_btn.config(state="normal", text="B.原单元格翻译")
        self.add_col_btn.config(state="normal", text="A.增一列老料号")
        self.select_btn.config(state="normal")
        messagebox.showerror("翻译失败", f"发生错误：\n{err_msg}")


def main():
    root = tk.Tk()
    TranslatorApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
